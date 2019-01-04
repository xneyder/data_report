from openpyxl import load_workbook

wb = load_workbook('templates/AVG_DATA_ANALISYS_TEMPLATE.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['B6'] = 0
ws['C6'] = 1

#ws['A3'] = 'Marry'
#ws['B3'] = 29

# Save the file
wb.save("sample.xlsx")
