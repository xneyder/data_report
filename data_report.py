#Daniel Jaramillo
#May 17 2018
#This script makes a list of all the tables in the library list given 
#Then per each table counts the number of records per day
#The result will be an excel with a tab per library and marks in red if a day has less records than the average
import sys
import os
import logging
import cx_Oracle
import base64
from logging.handlers import TimedRotatingFileHandler
from openpyxl import load_workbook
from datetime import timedelta,date,datetime
from threading import Thread


if len(sys.argv) < 3:
	print('Usage {script} [start_date] [end_date]'.format(script=sys.argv[0]))
	print('Example {script} "04/16/2018" "05/16/2018"'.format(script=sys.argv[0]))
	quit()

start_date=sys.argv[1]
end_date=sys.argv[2]

libraries=['ALU_ENUM_SPP','ALU_IPNE_SPP','AMD_APPL_SPP','BLU_IPAM_SPP','BMC_CMAS_SPP','BRIX_EXFO_FPP','BRW_AAA_SPP','CIS_ENV_SPP','CIS_IPSEC_SPP','CIS_IPSLA_SPP','CIS_IRONSMS_SPP','CIS_NTFLW_NPP_CDR','CIS_NTFLW_NPP_PM','CIS_QOS_SPP','ERI_CSCF_FPP','ERI_MGCF_FPP','ERI_MGW_FPP','ERI_MRS_XPP','ERI_SBG_FPP','EXFO_BRIXW_FPP','EXT_SW_SPP','FOR_FW_SPP','HP_HH3_SPP','JNX_COS_SPP','JNX_ENV_SPP','NOK_HLR_FPP','NOK_MDM_FPP','NOK_ONENDS_FPP','ORA_DRA_CPP','PAN_FW_SPP','RAD_LB_SPP','SNW_AVTL_SPP','STD_BGP_SPP','STD_IPIF_SPP','STD_IT_MONITORING_SPP']

DB_USER='psa'
DB_PASSWORD=base64.b64decode('dHRpcGFzcw==')
ORACLE_SID='IPHLXP'
DB_HOST='shhlxprd-scan:1521'
FORMATTER=logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
LOG_FILE=os.environ['LOG_DIR']+'/data_report.log'
DBL_PATH=os.environ['DVX2_IMP_DIR']+'/config/Dbl/'
datefield='DATETIME'
threads = 10


class ManagedDbConnection:
    def __init__(self, DB_USER,DB_PASSWORD,ORACLE_SID,DB_HOST):
        self.DB_USER = DB_USER
        self.DB_PASSWORD = DB_PASSWORD
        self.ORACLE_SID = ORACLE_SID
        self.DB_HOST = DB_HOST

    def __enter__(self):
        try:
            self.db = cx_Oracle.connect('{DB_USER}/{DB_PASSWORD}@{DB_HOST}/{ORACLE_SID}'.format(DB_USER=self.DB_USER,DB_PASSWORD=self.DB_PASSWORD,DB_HOST=self.DB_HOST,ORACLE_SID=self.ORACLE_SID), threaded=True)
        except cx_Oracle.DatabaseError as e:
            app_logger.error(e)
            quit()
        self.cursor = self.db.cursor()
        sqlplus_script="alter session set nls_date_format = 'DD-MON-YY HH24:MI'"
        try:
            self.cursor.execute(sqlplus_script)
        except cx_Oracle.DatabaseError as e:
            app_logger.error(e)
            app_logger.error(sqlplus_script[0:900])
            quit()
        return self.db

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.cursor:
            self.cursor.close()
        if self.db:
            self.db.close()

def get_console_handler():
    console_handler=logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(FORMATTER)
    return console_handler

def get_file_handler():
    file_handler=TimedRotatingFileHandler(LOG_FILE,when='midnight')
    file_handler.setFormatter(FORMATTER)
    return file_handler

def get_logger(logger_name):
    logger=logging.getLogger(logger_name)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(get_console_handler())
    logger.addHandler(get_file_handler())
    logger.propagate=False
    return logger

def build_ar_dict(keys,cursor):
    """
    Zip keys and values to a dictionary and build a list of them
    """
    data_dictionary=[]
    for row in filter(None,cursor):
        data_dictionary.append(dict(zip(keys, row)))
    return data_dictionary

def daterange(start_date,end_date):
	interval=int((end_date-start_date).days)+1
	for n in range(interval):
		yield start_date+timedelta(n)

def query_table(table_name,schema):
    """
    Process the table 
    """
    with ManagedDbConnection(DB_USER,DB_PASSWORD,ORACLE_SID,DB_HOST) as db:
		cursor=db.cursor()
		app_logger.info('Deleting records in AUDIT_DB.RECORDS_COUNT for {schema}.{table_name}'.format(schema=schema,table_name=table_name,))
		sqlplus_script="""
		DELETE FROM AUDIT_DB.RECORDS_COUNT 
		WHERE SCHEMAAA='{schema}' and TABLE_NAME='{table_name}'
		""".format(schema=schema,
			table_name=table_name,
			)
		try:
			cursor.execute(sqlplus_script)
		except cx_Oracle.DatabaseError as e:
			app_logger.error(e)
			app_logger.error(sqlplus_script)
			return
		db.commit()

		app_logger.info('counting the number of records for {schema}.{table_name}'.format(schema=schema,table_name=table_name,))
		sqlplus_script="""
		INSERT INTO AUDIT_DB.RECORDS_COUNT
		SELECT '{schema}','{table_name}',TRUNC({datefield}),COUNT(*)
		from {schema}.{table_name}
		where DATETIME>=TO_DATE('{start_date}','MM/DD/YYYY')
		and DATETIME<TO_DATE('{end_date}','MM/DD/YYYY')+1
		group by TRUNC({datefield})
		""".format(schema=schema,
			table_name=table_name,
			datefield=datefield,
			start_date=start_date,
			end_date=end_date,
			)
		try:
			cursor.execute(sqlplus_script)
		except cx_Oracle.DatabaseError as e:
			app_logger.error(e)
			app_logger.error(sqlplus_script)
			return
		db.commit()

def process_library(library):
	app_logger.info('Getting table list for {library}'.format(library=library))
	dbl_file=DBL_PATH+'/'+library+'.dbl'
	if not os.path.exists(dbl_file):
		app_logger.error('{dbl_file} does not exist'.format(dbl_file=dbl_file))
		return
	
	if library not in wb:
		app_logger.error('sheet {library} not found in template file'.format(library=library))
		return

	table_list=[]
	schema=''
	with open(dbl_file) as file:
		filedata=file.read()
		for line in filedata.split('\n'):
			if line.startswith('TargetTable='):
				table_list.append(line.split('=')[1])
			if line.startswith('DBProfile='):
				schema=line.split('=')[1]

	if not table_list:
		app_logger.error('No tables found for {library}'.format(library=library))
		return

	threads=[]
	for table_name in table_list:
		worker = Thread(target=query_table, args=(table_name,schema,))
		threads.append(worker)

	for worker in threads:
	    worker.setDaemon(True)
	    worker.start()

	for worker in threads:
		worker.join()

	ws = wb[library]
	for idx, v in enumerate(date_range_arr, 3): 
		ws.cell(row=2, column=idx, value=v) 

	for idx_table,table_name in enumerate(table_list,3):
	    with ManagedDbConnection(DB_USER,DB_PASSWORD,ORACLE_SID,DB_HOST) as db:
			cursor=db.cursor()
			sqlplus_script="""
			SELECT DATETIME,RECORDS
			from AUDIT_DB.RECORDS_COUNT
			where SCHEMAAA='{schema}' and TABLE_NAME='{table_name}'
			""".format(schema=schema,
				table_name=table_name,
				)
			try:
				cursor.execute(sqlplus_script)
			except cx_Oracle.DatabaseError as e:
				app_logger.error(e)
				app_logger.error(sqlplus_script)
				return

			db_records={}
			for row in filter(None,cursor):
				db_records[row[0].strftime('%m/%d/%Y')]=row[1]

			if len(db_records.values()) > 0:
				average=sum(db_records.values())/len(db_records.values())
			else:
				average=0

			xlsx_records=[table_name,average]
			for sdate in date_range_arr:
				if sdate not in db_records:
					db_records[sdate]=0
				xlsx_records.append(db_records[sdate])

			for idx, v in enumerate(xlsx_records, 1): 
				if v < average:
					v='('+str(v)+')'
				ws.cell(row=idx_table, column=idx, value=v) 
	app_logger.info('writing to data_report.xlsx for {library}'.format(library=library))
	wb.save("data_report.xlsx")

app_logger=get_logger('data_report')
app_logger.info('Starting the data report process')


app_logger.info('loading template AVG_DATA_ANALISYS_TEMPLATE.xlsx')
wb = load_workbook('templates/AVG_DATA_ANALISYS_TEMPLATE.xlsx')

date_range_arr=[]
dstart_date=datetime.strptime(start_date, '%m/%d/%Y')
dend_date=datetime.strptime(end_date, '%m/%d/%Y')
for tdate in daterange(dstart_date,dend_date):
	sdate=tdate.strftime('%m/%d/%Y')
	date_range_arr.append(sdate)

threads=[]
for library in libraries:
	worker = Thread(target=process_library, args=(library,))
	threads.append(worker)

for worker in threads:
    worker.setDaemon(True)
    worker.start()

for worker in threads:
	worker.join()
	
# wb.save("data_report.xlsx")
app_logger.info('File data_report.xlsx created')
